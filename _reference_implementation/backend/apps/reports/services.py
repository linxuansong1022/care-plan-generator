"""
Report service for generating export files.
Supports CSV and Excel (XLSX) formats.
"""

import csv
import io
from datetime import date, datetime
from typing import List, Optional, Literal

from django.db.models import Count, F
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.care_plans.models import CarePlan
from apps.orders.models import Order
from apps.patients.models import Patient
from apps.providers.models import Provider


class ReportService:
    """Service for generating export reports."""
    
    def export_orders(
        self,
        format: Literal["csv", "xlsx"] = "csv",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        status: Optional[str] = None,
        provider_npi: Optional[str] = None,
    ) -> tuple:
        """
        Export orders report.

        Returns:
            Tuple of (file_bytes, filename, content_type)
        """
        # Build query
        queryset = Order.objects.select_related(
            "patient", "provider"
        ).prefetch_related("care_plan")

        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)
        if status:
            queryset = queryset.filter(status=status)
        if provider_npi:
            queryset = queryset.filter(provider__npi=provider_npi)

        queryset = queryset.order_by("-created_at")

        # Prepare data
        headers = [
            "Order ID",
            "Created Date",
            "Status",
            "Patient MRN",
            "Patient Name",
            "Primary Diagnosis",
            "Medication",
            "Provider NPI",
            "Provider Name",
            "Care Plan Generated",
            "Care Plan Date",
        ]

        rows = []
        for order in queryset:
            has_care_plan = hasattr(order, "care_plan") and order.care_plan is not None
            rows.append([
                str(order.id),
                order.created_at.strftime("%Y-%m-%d %H:%M"),
                order.status,
                order.patient.mrn,
                f"{order.patient.first_name} {order.patient.last_name}",
                order.patient.primary_diagnosis_code,
                order.medication_name,
                order.provider.npi,
                order.provider.name,
                "Yes" if has_care_plan else "No",
                order.care_plan.generated_at.strftime("%Y-%m-%d %H:%M") if has_care_plan else "",
            ])

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if format == "csv":
            return (
                self._generate_csv(headers, rows),
                f"orders_export_{timestamp}.csv",
                "text/csv",
            )
        else:
            return (
                self._generate_xlsx(headers, rows, "Orders"),
                f"orders_export_{timestamp}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    def export_all_orders_with_care_plans(self) -> tuple:
        """
        Export all orders with care plan content for pharma reporting.

        CSV includes:
        - order_id, order_date
        - patient_mrn, patient_first_name, patient_last_name, patient_date_of_birth
        - provider_npi, provider_name
        - medication_name, primary_diagnosis_code
        - care_plan_status, care_plan_content

        Returns:
            Tuple of (file_bytes, filename, content_type)
        """
        queryset = Order.objects.select_related(
            "patient", "provider"
        ).prefetch_related("care_plan").order_by("-created_at")

        headers = [
            "order_id",
            "order_date",
            "patient_mrn",
            "patient_first_name",
            "patient_last_name",
            "patient_date_of_birth",
            "provider_npi",
            "provider_name",
            "medication_name",
            "primary_diagnosis_code",
            "care_plan_status",
            "care_plan_content",
        ]

        rows = []
        for order in queryset:
            has_care_plan = hasattr(order, "care_plan") and order.care_plan is not None

            # Determine care plan status
            if has_care_plan:
                care_plan_status = "completed"
                care_plan_content = order.care_plan.content or ""
            elif order.status == "failed":
                care_plan_status = "failed"
                care_plan_content = ""
            elif order.status == "processing":
                care_plan_status = "processing"
                care_plan_content = ""
            else:
                care_plan_status = "pending"
                care_plan_content = ""

            rows.append([
                str(order.id),
                order.created_at.strftime("%Y-%m-%d"),
                order.patient.mrn,
                order.patient.first_name,
                order.patient.last_name,
                str(order.patient.date_of_birth) if order.patient.date_of_birth else "",
                order.provider.npi,
                order.provider.name,
                order.medication_name,
                order.patient.primary_diagnosis_code,
                care_plan_status,
                care_plan_content,
            ])

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        return (
            self._generate_csv(headers, rows),
            f"orders_care_plans_export_{timestamp}.csv",
            "text/csv",
        )
    
    def export_provider_report(
        self,
        format: Literal["csv", "xlsx"] = "xlsx",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> tuple:
        """
        Export provider summary report (for pharma reporting).
        
        Groups orders by provider with aggregated statistics.
        """
        # Build query with aggregations
        queryset = Provider.objects.annotate(
            total_orders=Count("orders"),
            unique_patients=Count("orders__patient", distinct=True),
        ).filter(total_orders__gt=0)
        
        if start_date or end_date:
            order_filter = {}
            if start_date:
                order_filter["orders__created_at__date__gte"] = start_date
            if end_date:
                order_filter["orders__created_at__date__lte"] = end_date
            queryset = queryset.filter(**order_filter)
        
        queryset = queryset.order_by("-total_orders")
        
        # Prepare data
        headers = [
            "Provider NPI",
            "Provider Name",
            "Total Orders",
            "Unique Patients",
            "Completed Care Plans",
            "Completion Rate (%)",
        ]
        
        rows = []
        for provider in queryset:
            # Count completed care plans
            completed = CarePlan.objects.filter(order__provider=provider).count()
            completion_rate = (completed / provider.total_orders * 100) if provider.total_orders > 0 else 0
            
            rows.append([
                provider.npi,
                provider.name,
                provider.total_orders,
                provider.unique_patients,
                completed,
                f"{completion_rate:.1f}",
            ])
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format == "csv":
            return (
                self._generate_csv(headers, rows),
                f"provider_report_{timestamp}.csv",
                "text/csv",
            )
        else:
            return (
                self._generate_xlsx(headers, rows, "Provider Summary"),
                f"provider_report_{timestamp}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    
    def export_patient_history(
        self,
        patient_mrn: str,
        format: Literal["csv", "xlsx"] = "xlsx",
    ) -> tuple:
        """
        Export complete history for a single patient.
        """
        try:
            patient = Patient.objects.get(mrn=patient_mrn)
        except Patient.DoesNotExist:
            raise ValueError(f"Patient not found: {patient_mrn}")
        
        orders = Order.objects.filter(
            patient=patient
        ).select_related("provider").prefetch_related("care_plan").order_by("-created_at")
        
        headers = [
            "Order Date",
            "Medication",
            "Provider",
            "Provider NPI",
            "Status",
            "Care Plan Generated",
        ]
        
        rows = []
        for order in orders:
            has_care_plan = hasattr(order, "care_plan") and order.care_plan is not None
            rows.append([
                order.created_at.strftime("%Y-%m-%d"),
                order.medication_name,
                order.provider.name,
                order.provider.npi,
                order.status,
                "Yes" if has_care_plan else "No",
            ])
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"patient_{patient_mrn}_history_{timestamp}"
        
        if format == "csv":
            return (
                self._generate_csv(headers, rows),
                f"{filename}.csv",
                "text/csv",
            )
        else:
            return (
                self._generate_xlsx(
                    headers,
                    rows,
                    f"Patient {patient.first_name} {patient.last_name}",
                    patient_info={
                        "MRN": patient.mrn,
                        "Name": f"{patient.first_name} {patient.last_name}",
                        "DOB": str(patient.date_of_birth) if patient.date_of_birth else "N/A",
                        "Primary Diagnosis": patient.primary_diagnosis_code,
                    },
                ),
                f"{filename}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    
    def export_medication_summary(
        self,
        format: Literal["csv", "xlsx"] = "xlsx",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> tuple:
        """
        Export medication summary report.
        Groups orders by medication with counts.
        """
        queryset = Order.objects.values("medication_name").annotate(
            total_orders=Count("id"),
            unique_patients=Count("patient", distinct=True),
            unique_providers=Count("provider", distinct=True),
        )
        
        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)
        
        queryset = queryset.order_by("-total_orders")
        
        headers = [
            "Medication",
            "Total Orders",
            "Unique Patients",
            "Unique Providers",
        ]
        
        rows = [
            [
                item["medication_name"],
                item["total_orders"],
                item["unique_patients"],
                item["unique_providers"],
            ]
            for item in queryset
        ]
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format == "csv":
            return (
                self._generate_csv(headers, rows),
                f"medication_summary_{timestamp}.csv",
                "text/csv",
            )
        else:
            return (
                self._generate_xlsx(headers, rows, "Medication Summary"),
                f"medication_summary_{timestamp}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    
    def _generate_csv(self, headers: List[str], rows: List[List]) -> bytes:
        """Generate CSV file."""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        return output.getvalue().encode("utf-8")
    
    def _generate_xlsx(
        self,
        headers: List[str],
        rows: List[List],
        sheet_name: str = "Report",
        patient_info: Optional[dict] = None,
    ) -> bytes:
        """Generate Excel file with formatting."""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name limit
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        start_row = 1
        
        # Add patient info header if provided
        if patient_info:
            for key, value in patient_info.items():
                ws.cell(row=start_row, column=1, value=key).font = Font(bold=True)
                ws.cell(row=start_row, column=2, value=value)
                start_row += 1
            start_row += 1  # Empty row before data
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Add data rows
        for row_idx, row in enumerate(rows, start_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in rows:
                if col_idx <= len(row):
                    max_length = max(max_length, len(str(row[col_idx - 1])))
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
